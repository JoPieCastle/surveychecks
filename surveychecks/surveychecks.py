import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Color
from pathlib import Path
from collections import OrderedDict

from surveychecks.helper.docReader import docReader
from surveychecks.helper.parser import parser


class surveychecks:
    def __init__(self, dataframe, wordDocumentPath, infoIn="table"):
        self.dataframe = dataframe
        self.wordDocumentPath = Path(wordDocumentPath)

        self.doc = docReader(self.wordDocumentPath, infoIn)
        self.text = self.doc.getText()
        self.pars = parser(self.text)
        self.varlist = self.pars.getVarInfo()

        self.varnameList = self.makeVarnameList()
        self.rangeList = self.makeRangeList()
        self.filterDic = self.makeFilterDic()
        self.expandedFilterDic = self.expandFilterDic()
        self.allFilterDicConditions = self.makeAllFilterDicConditions()

    def makeFilterDic(self):
        """creates the filter dictionary necessary for filterCheck

        Returns:
            dictionary: filter dictionary with keys as variables and values filterconditions to be read with pd.eval
        """
        filterDic = OrderedDict()
        for var in self.varlist:
            try:
                templist = var.split(";")
                if len(templist) == 2:
                    filterDic[self.pars.inputToFullString(templist[0], out="variable")] = self.pars.inputToFullString(
                        templist[1].lstrip()
                    )
            except:
                raise Exception(f'filterDic parsing error at "{var}"')

        return filterDic

    def expandFilterDic(self):  # TODO update docstring
        """expands filterdic to include all filterchecks of variables within the filter e.g.:
        E0 == 1, 2; EX == 1 & EY >= 1
        E1 == 1, 2; E0 == 1 | EP == 2
        E2 == 1, 2; E1 == 1 | EM == 1
        Filtercondition Evaluation expansion for E2 should include all Filterconditions of E1 and previous, such that
        Filtercondition E2: (E1 == 1 & ((E0 == 1 & (EX == 1 & EY >= 1)) | EP == 2)) | EM == 1
        """
        expandedFilterDic = OrderedDict()
        for key, value in self.filterDic.items():
            # extending the filterDictionary
            expandedFilterDic[key] = self.pars.filterEvalExtender(value, self.filterDic)

        return expandedFilterDic

    def makeAllFilterDicConditions(self):  # TODO update docstring

        expandedFilterDic = OrderedDict()
        for key, value in self.filterDic.items():
            expandedFilterCondition = f"({value})"
            filterQuestionListToCheck = self.pars.createVarNameList(value)
            filterQuestionListChecked = []

            while True:
                # look for filterconditions in new variables of previous filtercondition
                loopTempList = filterQuestionListToCheck.copy()
                for filterQuestion in loopTempList:
                    if filterQuestion in self.filterDic.keys():
                        expandedFilterCondition += f" & ({self.filterDic[filterQuestion]})"
                    filterQuestionListChecked.append(filterQuestion)
                    filterQuestionListToCheck.remove(filterQuestion)

                # parse out all variables in the new expanded filtercondition
                newQuestionListToCheck = self.pars.createVarNameList(expandedFilterCondition)

                # check if there are new variables
                for question in newQuestionListToCheck:
                    if question not in filterQuestionListChecked:
                        filterQuestionListToCheck.append(question)

                # break loop if no new variables to be checked
                if len(filterQuestionListToCheck) == 0:
                    break

            expandedFilterDic[key] = self.pars.createVarNameList(expandedFilterCondition)

        return expandedFilterDic

    def makeRangeList(self):
        """creates the range dictionary necessary for rangeCheck

        Returns:
            list: range condition for each variable to be read with pd.eval
        """
        rangeList = []
        for var in self.varlist:
            try:
                templist = var.split(";")
                rangeList.append(self.pars.inputToFullString(templist[0]))
            except:
                raise Exception(f"rangeList parsing error {var}")

        return rangeList

    def makeVarnameList(self):
        """creates list of varnames for varCheck

        Returns:
            list: list consisting of varnames found in the document
        """
        varNameList = []
        for var in self.varlist:
            templist = var.split(";")
            varNameList.append(self.pars.inputToFullString(templist[0], out="variable"))

        return varNameList

    def varCheck(self, excelOut=False):  # TODO update docstring
        """Evaluates whether all variables in the file are in the data frame.

        Args:
            string: whether to look if variables are missing in doc file or dataframe

        Returns:
            list:
                [0] variables found in the data frame
                [1] variables not found in the data frame
        """
        dataVarnames = list(self.dataframe)

        bugCounter = 0
        includedVarnames = []
        excludedVarnamesDoc = []
        for varname in self.varnameList:
            if varname in dataVarnames:
                includedVarnames.append(varname)
            else:
                bugCounter += 1
                print(f'Variable "{varname}" not found in dataset')
                excludedVarnamesDoc.append(varname)

        print(f'"{len(includedVarnames)}" out of "{len(self.varnameList)}" doc variables included in dataframe')

        bugCounter = 0
        excludedVarnamesDF = []
        for varname in dataVarnames:
            if varname not in self.varnameList:
                bugCounter += 1
                print(f'Variable "{varname}" not found in doc')
                excludedVarnamesDF.append(varname)

        print(f'"{len(includedVarnames)}" out of "{len(dataVarnames)}" dataframe variables included in doc')

        if excelOut == True:
            allVars = includedVarnames + excludedVarnamesDoc + excludedVarnamesDF
            dfOut = pd.DataFrame(allVars, columns=["All Variables"])
            dfOut["Only in Document"] = dfOut["All Variables"].isin(excludedVarnamesDoc)
            dfOut["Only in Dataframe"] = dfOut["All Variables"].isin(excludedVarnamesDF)

            excelName = self.wordDocumentPath.with_name(f"VC_{self.wordDocumentPath.stem}.xlsx")
            with pd.ExcelWriter(excelName) as writer:
                dfOut.to_excel(writer, sheet_name=f"VarCheck")

            wb = openpyxl.load_workbook(excelName)
            ws = wb.active
            d = ws.cell(
                row=1,
                column=5,
                value=f'"{len(includedVarnames)}" out of "{len(self.varnameList)}" doc variables included in dataframe',
            )
            d.font = Font(color="008000" if len(excludedVarnamesDoc) == 0 else "FF0000", bold=True)

            d = ws.cell(
                row=2,
                column=5,
                value=f'"{len(includedVarnames)}" out of "{len(dataVarnames)}" dataframe variables included in doc',
            )
            d.font = Font(color="008000" if len(excludedVarnamesDF) == 0 else "FF0000", bold=True)

            wb.save(excelName)

        return {
            "In Document & Dataframe": includedVarnames,
            "In Dataframe only": excludedVarnamesDF,
            "In Document only": excludedVarnamesDoc,
        }

    def rangeCheck(self, checkType="unallowed", outList="failed", excelOut=False):  # TODO update docstring
        """Evaluates whether there are unallowed values in a variables range condition or
        if there are missings values in a variable given the specified range condition

        Args:
            checkType (string): whether to check for unallowed values or missing values
            outList (string): whether to return sucessfull checks or the failed checks

        checkType = 'unallowed' returns:
            list: variables with inconsistencies in rangeCheck unallowed
                [warn_num][0] = question
                [warn_num][1] = question and conditions for range
                [warn_num][2] = data frame for question by  range condition
                [warn_num][3] = full data frame filtered by range condition
        checkType = 'missing' returns:
            list: variables with inconsistencies in rangeCheck missing
                [warn_num][0] = question
                [warn_num][1] = question and conditions for range
                [warn_num][2] = values that were (not) found
        """
        bugCounter = -1
        checkCounter = 0
        outListFailed = []
        outListSuccess = []
        for var in self.rangeList:
            try:
                # creating varlist for potential problemlist output filtering
                checkCounter += 1
                singleVar = self.pars.inputToFullString(var, out="variable")  # filtering variable
                if checkType == "unallowed":
                    # filtering question by range condition
                    filt = self.dataframe.eval(var)
                    # reversed lookup, to see whether there is data although not in range
                    out = self.dataframe[~filt]
                    # check whether there are values in variable although not defined in range
                    if sum(pd.notnull(out[singleVar])) > 0:
                        bugCounter += 1
                        print(
                            f"{bugCounter}: Question '{singleVar}' given range condition '{var}' has unallowed values"
                        )
                        outListFailed.append([singleVar, var, out[singleVar].sort_values(), out])
                    else:
                        outListSuccess.append([singleVar, var, out[singleVar].sort_values(), out])

                elif checkType == "missing":
                    valueList = re.sub("\(|\)", "", var).split("|")
                    valueNotFoundList = []
                    valueFoundList = []

                    for val in valueList:
                        # querrying each value of rangelist seperately to assess if values are missing
                        valBool = self.dataframe.eval(val)

                        if sum(valBool) == 0:  # check wether the value exists
                            valueNotFoundList.append(self.pars.singleExpParse(val))
                        else:
                            valueFoundList.append(self.pars.singleExpParse(val))

                    if len(valueNotFoundList) != 0:
                        bugCounter += 1
                        print(f"{bugCounter}: Question '{singleVar}' given range condition '{var}' has missing values")
                        outListFailed.append([singleVar, var, valueNotFoundList])
                    else:
                        outListSuccess.append([singleVar, var, valueFoundList])

                else:
                    raise Exception('checkType not recognized -> only "unallowed" and "missing" are available')

            except:
                raise Exception(f"Failed at the following eval string: {var}")

        print(
            f"{bugCounter + 1} out of {checkCounter} variables showed inconsistencies running rangeCheck({checkType})"
        )

        if excelOut == True and checkType == "unallowed":
            excelName = self.wordDocumentPath.with_name(f"RC_{checkType}_{self.wordDocumentPath.stem}.xlsx")
            if len(outListFailed) > 0:
                with pd.ExcelWriter(excelName) as writer:
                    for warning in outListFailed:
                        warning[2].to_excel(writer, sheet_name=f"{warning[0]}")

                wb = openpyxl.load_workbook(excelName)
                for warning in outListFailed:
                    ws = wb[warning[0]]
                    d = ws.cell(row=1, column=3, value=f"Range Condition: {warning[1]}")
                    d.font = Font(color="FF0000", bold=True)
                wb.save(excelName)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active            
                d = ws.cell(row=1, column=1, value=f"rangeCheck unallowed did not find any inconsistencies")
                d.font = Font(bold=True)
                wb.save(excelName)

        if excelOut == True and checkType == "missing":
            excelName = self.wordDocumentPath.with_name(f"RC_{checkType}_{self.wordDocumentPath.stem}.xlsx")

            wb = openpyxl.Workbook()
            ws = wb.active
            d = ws.cell(row=1, column=1, value=f"Variable")
            d.font = Font(bold=True)
            d = ws.cell(row=1, column=2, value=f"Range Condition")
            d.font = Font(bold=True)
            d = ws.cell(row=1, column=3, value=f"Missing Values")
            d.font = Font(bold=True)

            rowCount = 2
            for warning in outListFailed:
                d = ws.cell(row=rowCount, column=1, value=f"{warning[0]}")
                d = ws.cell(row=rowCount, column=2, value=f"{warning[1]}")
                d = ws.cell(row=rowCount, column=3, value=f'{", ".join(warning[2])}')
                rowCount += 1
            wb.save(excelName)

        if outList == "failed":
            return outListFailed
        else:
            return outListSuccess

    def filterCheck(self, filterMissVal, expandedCheck=False, checkType="unallowed", excelOut=False):
        # TODO update docstring
        """Evaluates whether there are unallowed values in a filterfollowquestion or
        if there are missings in filterfollowquestion even though a filtercondition is true

        Args:
            checkType (string): whether one wants to check for unallowed values or missing values

        Returns:
            list: variables where the filterCheck showed inconsistencies
                [warn_num][0] = filter follow question
                [warn_num][1] = filter questions and conditions for filter follow question
                [warn_num][2] = data frame for filter and filter follow question filtered by filter condition(s)
                [warn_num][3] = full data frame filtered by filter condition(s)
        """
        bugCounter = -1
        checkCounter = 0
        outList = []
        if expandedCheck:
            filterDic = self.expandedFilterDic
        else:
            filterDic = self.filterDic

        for key, value in filterDic.items():
            try:
                checkCounter += 1
                # creating varlist for potential problemlist output filtering
                varlist = re.findall(r"[(&|;]+\s*[(&|;]*(.*?)\s[!=><]=\s", "|" + str(value))  # filtering variables
                varlist = list(dict.fromkeys(varlist))  # removing duplicates
                relevant_vars = [key] + varlist

                # filtering filterfollowquestion by filter condition
                filt = self.dataframe.eval(value)
                
                if checkType == "unallowed":
                    # reversed lookup, to see whether there is data although filter condition not satisfied
                    out = self.dataframe[~filt]
                                    
                    # check whether there are values in filterfollowquestion although condition is not satisfied
                    #if sum(pd.isnotnull(out[key])) > 0:
                    if (out[key] == filterMissVal).sum() != len(out[key]):
                        bugCounter += 1
                        print(
                            f"{bugCounter}: Filter follow question '{key}' given filter condition'{value}' has unallowed values"
                        )
                        outList.append(
                            [
                                key,
                                value,
                                out[relevant_vars].sort_values(by=relevant_vars),
                                out.sort_values(by=relevant_vars),
                            ]
                        )

                elif checkType == "missing":
                    # finding rows fitting filtercondition
                    out = self.dataframe[filt]
                    # check that there are no missings in a filterfollowquestion given that filter condition is satisfied
                    #if sum(pd.isnull(out[key])) > 0:
                    if (out[key] == filterMissVal).sum() > 0:
                        bugCounter += 1
                        print(
                            f"{bugCounter}: Filter follow question '{key}' given filter condition '{value}' has missing values"
                        )
                        outList.append(
                            [
                                key,
                                value,
                                out[relevant_vars].sort_values(by=relevant_vars, na_position='first'),
                                out.sort_values(by=relevant_vars),
                            ]
                        )
                else:
                    raise Exception('checkType not recognized -> only "unallowed" and "missing" are available')
            except:
                raise Exception(f'Failed at the following evaluation: variable "{key}" with filter condition "{value}"')

        print(
            f"{bugCounter +1 } out of {checkCounter} variables showed inconsistencies running filterCheck({checkType})"
        )

        if excelOut == True:
            excelName = self.wordDocumentPath.with_name(f"FC_{checkType}_{self.wordDocumentPath.stem}.xlsx")
            if len(outList) > 0:
                with pd.ExcelWriter(excelName) as writer:
                    for warning in outList:
                        warning[2].to_excel(writer, sheet_name=f"{warning[0]}")
                wb = openpyxl.load_workbook(excelName)
                for warning in outList:
                    ws = wb[warning[0]]
                    d = ws.cell(
                        row=1,
                        column=len(warning[2].columns) + 2,
                        value=f"Filtercondition: {warning[1]}",
                    )
                    d.font = Font(color="FF0000", bold=True)
                wb.save(excelName)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active            
                d = ws.cell(row=1, column=1, value=f"filterCheck {checkType} did not find any inconsistencies")
                d.font = Font(bold=True)
                wb.save(excelName)

        return outList
